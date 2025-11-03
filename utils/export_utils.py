"""
Export Utilities
Export data to various formats (CSV, JSON, Excel)
"""
import csv
import json
from io import StringIO, BytesIO
from datetime import datetime
from pathlib import Path

from utils.logger import get_logger

logger = get_logger(__name__)

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logger.warning("openpyxl not installed. Excel export disabled.")


def export_to_json(data, filename=None):
    """
    Export data to JSON
    
    Args:
        data: List of dictionaries or single dictionary
        filename: Optional filename to save to
        
    Returns:
        str: JSON string or path to saved file
    """
    try:
        json_str = json.dumps(data, ensure_ascii=False, indent=2, default=str)
        
        if filename:
            filepath = Path('exports') / filename
            filepath.parent.mkdir(exist_ok=True)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(json_str)
            
            logger.info(f"Exported to JSON: {filepath}")
            return str(filepath)
        
        return json_str
        
    except Exception as e:
        logger.error(f"Error exporting to JSON: {e}")
        raise


def export_to_csv(data, filename=None):
    """
    Export data to CSV
    
    Args:
        data: List of dictionaries
        filename: Optional filename to save to
        
    Returns:
        str: CSV string or path to saved file
    """
    try:
        if not data:
            return "" if not filename else None
        
        # Get all unique keys from all dictionaries
        fieldnames = set()
        for item in data:
            if isinstance(item, dict):
                fieldnames.update(item.keys())
        
        fieldnames = sorted(list(fieldnames))
        
        # Create CSV
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        
        for item in data:
            # Convert nested objects to JSON strings
            row = {}
            for key, value in item.items():
                if isinstance(value, (dict, list)):
                    row[key] = json.dumps(value, ensure_ascii=False)
                else:
                    row[key] = value
            writer.writerow(row)
        
        csv_str = output.getvalue()
        
        if filename:
            filepath = Path('exports') / filename
            filepath.parent.mkdir(exist_ok=True)
            
            with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                f.write(csv_str)
            
            logger.info(f"Exported to CSV: {filepath}")
            return str(filepath)
        
        return csv_str
        
    except Exception as e:
        logger.error(f"Error exporting to CSV: {e}")
        raise


def export_to_excel(data, filename=None, sheet_name='Data'):
    """
    Export data to Excel (requires openpyxl)
    
    Args:
        data: List of dictionaries
        filename: Optional filename to save to
        sheet_name: Name of the worksheet
        
    Returns:
        BytesIO or path to saved file
    """
    if not EXCEL_SUPPORT:
        raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
    
    try:
        if not data:
            return None
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Get headers
        fieldnames = set()
        for item in data:
            if isinstance(item, dict):
                fieldnames.update(item.keys())
        
        fieldnames = sorted(list(fieldnames))
        
        # Write headers
        ws.append(fieldnames)
        
        # Write data
        for item in data:
            row = []
            for field in fieldnames:
                value = item.get(field, '')
                # Convert nested objects to JSON strings
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                row.append(value)
            ws.append(row)
        
        if filename:
            filepath = Path('exports') / filename
            filepath.parent.mkdir(exist_ok=True)
            
            wb.save(filepath)
            logger.info(f"Exported to Excel: {filepath}")
            return str(filepath)
        
        # Return BytesIO for download
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise


def export_invoices(invoices, format='json', filename=None):
    """
    Export invoices in specified format
    
    Args:
        invoices: List of invoice dictionaries
        format: 'json', 'csv', or 'excel'
        filename: Optional filename
        
    Returns:
        str or BytesIO: Exported data
    """
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"invoices_{timestamp}.{format}"
    
    if format == 'json':
        return export_to_json(invoices, filename)
    elif format == 'csv':
        return export_to_csv(invoices, filename)
    elif format == 'excel':
        return export_to_excel(invoices, filename, sheet_name='Invoices')
    else:
        raise ValueError(f"Unsupported format: {format}")


def export_forecasts(forecasts, format='json', filename=None):
    """
    Export forecasts in specified format
    
    Args:
        forecasts: List of forecast dictionaries
        format: 'json', 'csv', or 'excel'
        filename: Optional filename
        
    Returns:
        str or BytesIO: Exported data
    """
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"forecasts_{timestamp}.{format}"
    
    if format == 'json':
        return export_to_json(forecasts, filename)
    elif format == 'csv':
        return export_to_csv(forecasts, filename)
    elif format == 'excel':
        return export_to_excel(forecasts, filename, sheet_name='Forecasts')
    else:
        raise ValueError(f"Unsupported format: {format}")


def create_summary_report(invoices, forecasts, statistics, filename=None):
    """
    Create comprehensive report with multiple sheets (Excel only)
    
    Args:
        invoices: List of invoices
        forecasts: List of forecasts
        statistics: Statistics dictionary
        filename: Optional filename
        
    Returns:
        str: Path to saved file
    """
    if not EXCEL_SUPPORT:
        raise ImportError("openpyxl not installed")
    
    try:
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet('Summary')
        ws_summary.append(['Metric', 'Value'])
        for key, value in statistics.items():
            ws_summary.append([key.replace('_', ' ').title(), value])
        
        # Invoices sheet
        if invoices:
            ws_invoices = wb.create_sheet('Invoices')
            fieldnames = sorted(list(set().union(*[item.keys() for item in invoices])))
            ws_invoices.append(fieldnames)
            
            for item in invoices:
                row = []
                for field in fieldnames:
                    value = item.get(field, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    row.append(value)
                ws_invoices.append(row)
        
        # Forecasts sheet
        if forecasts:
            ws_forecasts = wb.create_sheet('Forecasts')
            fieldnames = sorted(list(set().union(*[item.keys() for item in forecasts])))
            ws_forecasts.append(fieldnames)
            
            for item in forecasts:
                row = []
                for field in fieldnames:
                    value = item.get(field, '')
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    row.append(value)
                ws_forecasts.append(row)
        
        # Save
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"summary_report_{timestamp}.xlsx"
        
        filepath = Path('exports') / filename
        filepath.parent.mkdir(exist_ok=True)
        
        wb.save(filepath)
        logger.info(f"Created summary report: {filepath}")
        return str(filepath)
        
    except Exception as e:
        logger.error(f"Error creating summary report: {e}")
        raise
